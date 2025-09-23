import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import load_workbook

def Decimals(df, columns):
    for col in columns:
        if col in df.columns:
            df[col] = df[col].map(lambda x: f"{float(x):.2f}" if pd.notnull(x) else "")
    return df


class DrippyKit:
    # ---------------------------- STYLE DEFINITIONS ----------------------------
    HeaderColor = PatternFill(start_color="4B1395", end_color="4B1395", fill_type="solid")
    SubHeaderColor = PatternFill(start_color="CBCEDF", end_color="CBCEDF", fill_type="solid")
    BodyColor = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    Border_IttyBitty = Side(style="thin", color="000000")
    Border_SheThicc = Side(style="thick", color="000000")
    Border_DoubleBottom = Side(style="double", color="000000")

    MainHeaderFont = Font(name="Aptos Display", bold=True, color="FFFFFF", size=14)
    HeaderFont = Font(name="Aptos Narrow", bold=True, color="FFFFFF", size=10)
    SubHeaderFont = Font(name="Aptos Narrow", bold=True, color="000000", size=10)
    BodyFont = Font(name="Aptos Narrow", bold=False, color="000000", size=10)
    BodyBoldFont = Font(name="Aptos Narrow", bold=True, color="000000", size=10)

    Middle = Alignment(horizontal="center", vertical="center", wrap_text=True)
    Lefty = Alignment(horizontal="left", vertical="center", indent=1)
    Righty = Alignment(horizontal="right", vertical="center", indent=1)

    HeaderStyle = {
        "font": HeaderFont,
        "alignment": Middle,
        "fill": HeaderColor,
        "border": Border(left=Border_IttyBitty, 
                        right=Border_IttyBitty, 
                        top=Border_IttyBitty, 
                        bottom=Border_DoubleBottom)
                    }

    ItemStyle = {
        "font": BodyFont,
        "alignment": Middle,
        "fill": BodyColor,
        "border": Border(left=Border_IttyBitty, 
                        right=Border_IttyBitty, 
                        top=Border_IttyBitty, 
                        bottom=Border_IttyBitty)
                    }

    SubHeaderStyle = {
        "font": BodyBoldFont,
        "alignment": Righty,
        "fill": SubHeaderColor,
        "border": Border(left=Border_IttyBitty, 
                        right=Border_IttyBitty, 
                        top=Border_IttyBitty, 
                        bottom=Border_IttyBitty)
                    }

    # ---------------------------- INSTANCE SETUP ----------------------------
    def __init__(self, filepath, sheet_name):
        self.filepath = filepath
        self.sheet_name = sheet_name
        self.wb = load_workbook(filepath)
        self.ws = self.wb[sheet_name]

    # ---------------------------- STYLING HELPERS ----------------------------
    @staticmethod
    def GetTheDrip(cell, font=None, alignment=None, fill=None, border=None):
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment
        if fill:
            cell.fill = fill
        if border:
            cell.border = border

    @staticmethod
    def Thicc(ws, FirstRow, LastRow, FirstColumn, LastColumn):
        """Apply a thick border around the given cell range."""
        for row in ws.iter_rows(min_row=FirstRow, max_row=LastRow, min_col=FirstColumn, max_col=LastColumn):
            for cell in row: # Start with the existing border
                left   = cell.border.left
                right  = cell.border.right
                top    = cell.border.top
                bottom = cell.border.bottom
                if cell.row == FirstRow: # Replace with thick sides where appropriate
                    top = DrippyKit.Border_SheThicc
                if cell.row == LastRow:
                    bottom = DrippyKit.Border_SheThicc
                if cell.column == FirstColumn:
                    left = DrippyKit.Border_SheThicc
                if cell.column == LastColumn:
                    right = DrippyKit.Border_SheThicc
                cell.border = Border(left=left, right=right, top=top, bottom=bottom) # Assign the new border back to the cell


    # ---------------------------- WORKBOOK STYLING ----------------------------
    def HeaderLewk(self):
        self.ws.row_dimensions[1].height = 45
        HeaderRow = next(self.ws.iter_rows(min_row=1, max_row=1))
        for cell in HeaderRow:
            self.GetTheDrip(cell, **self.HeaderStyle)

    def ItemsLewk(self):
        FirstItemsRow = 2
        LastRow = self.ws.max_row
        LastColumn = self.ws.max_column
        for row in range(FirstItemsRow, LastRow + 1):
            self.ws.row_dimensions[row].height = 15
            for col in range(1, LastColumn + 1):
                cell = self.ws.cell(row, col)
                self.GetTheDrip(cell, **self.ItemStyle)

    def ThiccBorder(self):
        LastRow = self.ws.max_row
        LastColumn = self.ws.max_column
        self.Thicc(self.ws, 1, LastRow, 1, LastColumn)

    def ColumnWidths(self):
        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]:
            self.ws.column_dimensions[col].width = 16

    def drip(self):
        self.HeaderLewk()
        self.ItemsLewk()
        self.ColumnWidths()
        self.ThiccBorder()
        self.wb.save(self.filepath)