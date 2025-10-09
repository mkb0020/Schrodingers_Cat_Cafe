import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from io import BytesIO
import pandas as pd
from RegressionCat import PurrfectRegression
import os
import tempfile
import atexit


class TipToe:
    # ---------------------------- STYLE DEFINITIONS ----------------------------
    PurpleyBlue = "4B1395"
    DarkishGrey = "75788B"
    LightBlueGrey = "CBCEDF"
    White = "FFFFFF"
    Black = "000000"
    SuperLightPink = "FFEFFF"
    SuperLightBlue = "E2E9FE"
    SuperLightPurp = "F0E1FF"

    PurpleFill = PatternFill(start_color=PurpleyBlue, end_color=PurpleyBlue, fill_type="solid")
    GreyFill = PatternFill(start_color=DarkishGrey, end_color=DarkishGrey, fill_type="solid")
    WhiteFill = PatternFill(start_color=White, end_color=White, fill_type="solid")
    BlueGreyFill = PatternFill(start_color=LightBlueGrey, end_color=LightBlueGrey, fill_type="solid")
    LightPinkFill = PatternFill(start_color=SuperLightPink, end_color=SuperLightPink, fill_type="solid")
    LightBlueFill = PatternFill(start_color=SuperLightBlue, end_color=SuperLightBlue, fill_type="solid")
    LightPurpleFill = PatternFill(start_color=SuperLightPurp, end_color=SuperLightPurp, fill_type="solid")

    BigBoldWhite = Font(name="Aptos Display", bold=True, color=White, size=14)
    BoldWhite = Font(name="Aptos Narrow", bold=True, color=White, size=11)
    BoldBlack = Font(name="Aptos Narrow", bold=True, color=Black, size=11)
    RegularBlack = Font(name="Aptos Narrow", bold=False, color=Black, size=11)

    Middle = Alignment(horizontal="center", vertical="center", wrap_text=True)
    Lefty = Alignment(horizontal="left", vertical="center", indent=1)
    Righty = Alignment(horizontal="right", vertical="center", indent=1)

    IttyBitty = Side(style="thin", color=Black)
    SheThicc = Side(style="thick", color=Black)
    Double = Side(style="double", color=Black)
    DoubleBottom = Border(bottom=Double)
    # ---------------------------- WALK-IN CLOSET / STYLES / Named after Riff Raff Albums ----------------------------
    PurpleIcon = { #BLUEISH-PURPLE BACKGROUND, BIG WHITE BOLLD FONT, CENTER ALIGNED, DOUBLE BOTTOM BORDER  + THICK SMALL RIGHT LEFT AND TOP BORDERS
        "font": BigBoldWhite,
        "alignment": Middle,
        "fill": PurpleFill,
        "border": Border(left=SheThicc, right=SheThicc, top=SheThicc, bottom=Double)
    }
    PurpleIconChopped = { #BLUEISH-PURPLE BACKGROUND, WHITE BOLLD FONT, CENTER ALIGNED, DOUBLE BOTTOM BORDER  + SMALL RIGHT LEFT AND TOP BORDERS
        "font": BoldWhite,
        "alignment": Middle,
        "fill": PurpleFill,
        "border": Border(left=IttyBitty, right=IttyBitty, top=IttyBitty, bottom=Double)
    }
    TheWhiteWest = { #DARK GREY BACKGROUND, WHITE BOLLD FONT, CENTER ALIGNED, DOUBLE BOTTOM BORDER + SMALL RIGHT LEFT AND TOP BORDERS
        "font": BoldWhite,
        "alignment": Middle,
        "fill": GreyFill,
        "border": Border(left=IttyBitty, right=IttyBitty, top=IttyBitty, bottom=Double)
    }
    TheWhiteWestChopped = { #DARK GREY BACKGROUND, WHITE BOLLD FONT, CENTER ALIGNED, DOUBLE BOTTOM BORDER ONLY
        "font": BoldWhite,
        "alignment": Middle,
        "fill": GreyFill,
        "border": Border(left=None, right=None, top=None, bottom=Double)
    }
    CoolBlueJewels = { #LIGHT BLUE-GREY BACKGROUND, BLACK BOLLD FONT, LEFT ALIGNED, SMALL BORDERS
        "font": BoldBlack,
        "alignment": Lefty,
        "fill": BlueGreyFill,
        "border": Border(left=IttyBitty, right=IttyBitty, top=IttyBitty, bottom=IttyBitty)
    }
    VanillaGorillaMids = { #WHITE BACKGROUND, BLACK REGULAR FONT, CENTER ALIGNED, SMALL BORDERS
        "font": RegularBlack,
        "alignment": Middle,
        "fill": WhiteFill,
        "border": Border(left=IttyBitty, right=IttyBitty, top=IttyBitty, bottom=IttyBitty)
    }
    VanillaGorillaLefts = { #WHITE BACKGROUND, BLACK REGULAR FONT, LEFT ALIGNED, SMALL BORDERS
        "font": RegularBlack,
        "alignment": Lefty,
        "fill": WhiteFill,
        "border": Border(left=IttyBitty, right=IttyBitty, top=IttyBitty, bottom=IttyBitty)
    }
    PinkPython = { #LIGHT PINK BACKGROUND, BLACK REGULAR FONT, CENTER ALIGNED, SMALL BORDERS
        "font": RegularBlack,
        "alignment": Middle,
        "fill": LightPinkFill,
        "border": Border(left=IttyBitty, right=IttyBitty, top=IttyBitty, bottom=IttyBitty)
    }
    AquaberryAquarius = { #LIGHT BLUE BACKGROUND, BLACK REGULAR FONT, CENTER ALIGNED, SMALL BORDERS
        "font": RegularBlack,
        "alignment": Middle,
        "fill": LightBlueFill,
        "border": Border(left=IttyBitty, right=IttyBitty, top=IttyBitty, bottom=IttyBitty)
    }
    LilacLightning = { #LIGHT PURPLE BACKGROUND, BLACK REGULAR FONT, CENTER ALIGNED, SMALL BORDERS
        "font": RegularBlack,
        "alignment": Middle,
        "fill": LightPurpleFill,
        "border": Border(left=IttyBitty, right=IttyBitty, top=IttyBitty, bottom=IttyBitty)
    }
    TurquoiseTornado = { #LIGHT BLUE-GREY BACKGROUND, BLACK REGULAR FONT, CENTER ALIGNED, NO BORDERS
        "font": RegularBlack,
        "alignment": Middle,
        "fill": BlueGreyFill
    }
    # ---------------------------- INIT ----------------------------
    def __init__(self, wb, ws):
        self.wb = wb
        self.ws = ws
    # ---------------------------- HELPERS ----------------------------
    @staticmethod
    def Flex(cell, font=None, alignment=None, fill=None, border=None): #APPLIES THE STYLE ABOVE TO CELLS
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment
        if fill:
            cell.fill = fill
        if border:
            cell.border = border

    def Thicc(self, FirstRow, LastRow, FirstColumn, LastColumn): #OUTLINE / APPLIES THICK OUTSIDE BORDER TO A RANGE OF CELLS
        for row in self.ws.iter_rows(min_row=FirstRow, max_row=LastRow,
                                     min_col=FirstColumn, max_col=LastColumn):
            for cell in row:
                left, right, top, bottom = cell.border.left, cell.border.right, cell.border.top, cell.border.bottom
                if cell.row == FirstRow:
                    top = TipToe.SheThicc
                if cell.row == LastRow:
                    bottom = TipToe.SheThicc
                if cell.column == FirstColumn:
                    left = TipToe.SheThicc
                if cell.column == LastColumn:
                    right = TipToe.SheThicc
                cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    def ColorBackGround(self, FirstRow, LastRow, FirstColumn, LastColumn): #COLORS THE BACKGROUND - USED IN THE TABS WITH PLOTS
        """Apply background color behind the plots."""
        for row in self.ws.iter_rows(min_row=FirstRow, max_row=LastRow,
                                     min_col=FirstColumn, max_col=LastColumn):
            for cell in row:
                cell.fill = TipToe.BlueGreyFill

    def SetColumnWidths(self, ThiccColumns, ThiccWidth, SpacerColumns, SpacerWidth): #SELF EXPLANITORY
        for col in ThiccColumns:
            self.ws.column_dimensions[col].width = ThiccWidth
        for col in SpacerColumns:
            self.ws.column_dimensions[col].width = SpacerWidth

    def Title(self, FirstRow): #APPLIES PURPLE ICON STYLE TO THE TITLE HEADER ON A TAB
        self.ws.row_dimensions[FirstRow].height = 25
        TitleRow = next(self.ws.iter_rows(min_row=FirstRow, max_row=FirstRow))
        for cell in TitleRow:
            TipToe.Flex(cell, **TipToe.PurpleIcon)

    def SubHeader(self, Row=2):  # FOR SCATTER AND RESULTS TABS - UNDER THE TITLE CELL
        self.ws.row_dimensions[Row].height = 20
        SubHeaderRow = next(self.ws.iter_rows(min_row=Row, max_row=Row))
        for cell in SubHeaderRow:
            TipToe.Flex(cell, **TipToe.TheWhiteWestChopped)

    # ---------------------------- WORKBOOK STYLING ----------------------------
    def InsightsDrip(self): #FIRST TAB - "INSIGHTS" - TITLE CELL = A1, CAT INSIGHTS SECTION = A3:E14, MODEL INSIGHTS = A16:A27
        FirstBorderRow = 1
        FirstColumn = 1
        LastRow = 27
        LastColumn = 5
        ThiccColumns = ["E"]
        ThiccWidth = 80
        SpacerColumns = ["B", "C", "D"]
        SpacerWidth = 17
        FirstFeatureRow = 3
        LastFeatureRow = 14
        FisrtMetricsRow = 16
        LastMetricsRow = 27
        self.Title(FirstBorderRow)
        self.SetColumnWidths(ThiccColumns, ThiccWidth, SpacerColumns, SpacerWidth)
        self.ws.column_dimensions["A"].width = 30
        self.Thicc(FirstBorderRow, LastRow, FirstColumn, LastColumn)
        self.Thicc(FirstFeatureRow, LastFeatureRow, FirstColumn, LastColumn)
        self.Thicc(FisrtMetricsRow, LastMetricsRow, FirstColumn, LastColumn)
        
    def MetricsDrip(self): #FIRST TAB - INSIGHTS - APPLIES VANILLA GORILLA MIDS TO  THE COEFFICIENTS COLUMNS AND THE METRICS COLUMNS (B-D) WHILE TAKING THE SPACERS INTO ACCOUNT
        ws = self.ws
        FirstRow1 = 5
        LastRow1 = 15
        FirstRow2 = 18
        LastRow2 = 28
        FirstColumn = 2
        LastColumn = 5
        for col in range(FirstColumn, LastColumn):
            for row in range(FirstRow1, LastRow1):
                cell = self.ws.cell(row, col)
                self.Flex(cell, **self.VanillaGorillaMids)
            for row in range(FirstRow2, LastRow2):
                cell = self.ws.cell(row, col)
                self.Flex(cell, **self.VanillaGorillaMids)

    def MetricsDripChopped(self): #FIRST TAB - INSIGHTS - APPLIES VANILLA GORILLA LEFTS TO  THE INSIGHTS COLUMN (E) WHILE TAKING THE SPACERS INTO ACCOUNT
        ws = self.ws
        FirstRow1 = 5
        LastRow1 = 15
        FirstRow2 = 18
        LastRow2 = 28
        FirstColumn = 5
        LastColumn = 5
        for row in range(FirstRow1, LastRow1):
            for col in range(FirstColumn, LastColumn):
                cell = self.ws.cell(row, col)
                self.Flex(cell, **self.VanillaGorillaLefts)
        for row in range(FirstRow2, LastRow2):
            for col in range(FirstColumn, LastColumn):
                cell = self.ws.cell(row, col)
                self.Flex(cell, **self.VanillaGorillaLefts)

    def DataDrip(self): #SECOND TAB - "DATA" - APPLIES COLUMN WIDTHS, STYLES HEADERS WITH PURPLE ICON CHOPPED, STYLES ROWS UNDER HEADERS WITH VANILLA GORILLA MIDS, AND APPLIES THICK OUTLINE
        ws = self.ws
        FirstBorderRow = 1
        FirstColumn = 1
        LastRow = self.ws.max_row
        LastColumn = self.ws.max_column
        ThiccColumns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P"]
        ThiccWidth = 15
        SpacerColumns = []
        SpacerWidth = 0
        self.ws.row_dimensions[1].height = 50
        HeaderRow = next(self.ws.iter_rows(min_row=1, max_row=1))
        for cell in HeaderRow:
            self.Flex(cell, **self.PurpleIconChopped)
        for row in range(FirstBorderRow+1, LastRow + 1):
            self.ws.row_dimensions[row].height = 15
            for col in range(FirstColumn, LastColumn + 1):
                cell = self.ws.cell(row, col)
                self.Flex(cell, **self.VanillaGorillaMids)
        self.SetColumnWidths(ThiccColumns, ThiccWidth, SpacerColumns, SpacerWidth)
        self.Thicc(FirstBorderRow, LastRow, FirstColumn, LastColumn)

    def PlotsDrip(self): #THIRD - SIXTH TABS - SCATTER PLOTS AND LINEAR REGRESSION PLOTS - APPLIES COLUMN WIDTHS, STYLES TITLE WITH PURPLE ICON, APPLIES THICK OUTLINE, APPLIES BACKGROUND COLOR BEHIND PLOT IMAGES, AND STYLES SUB-HEADERS WITH WHITE WEST
        ws = self.ws
        PlotTabs = ["SCATTER PLOTS", "MOOD RESULTS" "SASS RESULTS", "SURVIVAL RESULTS"]
        for tab in PlotTabs:
            tab = self.ws.title
            if tab == "SCATTER PLOTS":
                FirstBorderRow = 1
                FirstBGRow = 3
                FirstColumn = 1
                LastRow = 63
                LastColumn = 7
                ThiccColumns = ["B", "D", "F"]
                ThiccWidth = 50
                SpacerColumns = ["A", "C", "E", "G"]
                SpacerWidth = 3
            elif tab == "MOOD RESULTS":
                FirstBorderRow = 1
                FirstBGRow = 3
                FirstColumn = 1
                LastRow =48
                LastColumn = 9
                ThiccColumns = ["B", "D", "F", "H"]
                ThiccWidth = 38.5
                SpacerColumns = ["A", "C", "E", "G", "I"]
                SpacerWidth = 2
            elif tab == "SASS RESULTS":
                FirstBorderRow = 1
                FirstBGRow = 3
                FirstColumn = 1
                LastRow =23
                LastColumn = 5
                ThiccColumns = ["B", "D"]
                ThiccWidth = 75
                SpacerColumns = ["A", "C", "E"]
                SpacerWidth = 3
            elif tab == "SURVIVAL RESULTS":
                FirstBorderRow = 1
                FirstBGRow = 3
                FirstColumn = 1
                LastRow =23
                LastColumn = 3
                ThiccColumns = ["B"]
                ThiccWidth = 90
                SpacerColumns = ["A", "C"]
                SpacerWidth = 10
            self.Title(FirstBorderRow)
            self.SubHeader()
            self.SetColumnWidths(ThiccColumns, ThiccWidth, SpacerColumns, SpacerWidth)
            self.Thicc(FirstBorderRow, LastRow, FirstColumn, LastColumn)
            self.ColorBackGround(FirstBGRow, LastRow, FirstColumn, LastColumn)

    def Drip(self): #APPLY THE DRIP
        self.InsightsDrip()
        self.MetricsDrip()
        self.MetricsDripChopped()
        self.DataDrip()
        self.PlotsDrip()
        self.wb.save(self)

def CreateWB(OutputPath, TabNames): #HELPER TO CREATE WORKBOOK
    wb = Workbook()
    if "Sheet" in wb.sheetnames: #REMOVE DEFAULT SHEET
        del wb["Sheet"]
    for name in TabNames:
        wb.create_sheet(title=name)
    wb.save(OutputPath)
    return wb

class PurrfectWB:
    def __init__(self, df, OutputPath, ScatterPlots, RegressionPlots, Importances):
        self.df = df
        self.OutputPath = OutputPath
        self.wb = openpyxl.Workbook()
        DefaultTab = self.wb.active # REMOVE DEFAULT TAB
        self.wb.remove(DefaultTab)
        self.sheets = ["INSIGHTS", "DATA", "SCATTER PLOTS", 
                       "MOOD RESULTS", "SASS RESULTS", "SURVIVAL RESULTS"] #TABS
        self.ws = {}
        for name in self.sheets: # CREATE TABS
            self.ws[name] = self.wb.create_sheet(title=name)
        self.wb.save(self.OutputPath)         # SAVE IMMEDIATELY SO IT'S WRITTEN ONTO DISK
        self.ScatterPlots = ScatterPlots
        self.RegressionPlots = RegressionPlots
        self.Importances = Importances
        
    def SAVE(self): #SAVE AND REMOVE TEMPS
        self.wb.save(self.OutputPath)
        if hasattr(self.wb, "_image_streams"):
            self.wb._image_streams.clear()

    def GetWS(self, name): #SAFELY GET A TAB BY NAME
        if name in self.ws:
            return self.ws[name]
        raise ValueError(f"Worksheet '{name}' not found in {self.OutputPath}")
    
    def InsightsKitten(self, ws=None): #INSIGHTS TAB - FILL IN BASIC INFO, APPLY STYLING, INSERT COEFFICIENTS
        InsightsWS = self.GetWS("INSIGHTS")
        InsightsWS["A1"] = "CAFÉ ANALYSIS" #TITLE
        InsightsWS.merge_cells("A1:E1")
        InsightsWS.merge_cells("A2:E2") #SPACER

        Headers = [("A3", "FEATURE"), ("B3", "IMPORTANCE"), ("E3", "INSIGHTS")]
        for cell, text in Headers:
            InsightsWS[cell] = text
            TipToe.Flex(InsightsWS[cell], **TipToe.PurpleIconChopped)
        InsightsWS.merge_cells("B3:D3")

        for col, text in [("B4", "MOOD"), ("C4", "SASS"), ("D4", "SURVIVAL")]:
            InsightsWS[col] = text
            TipToe.Flex(InsightsWS[col], **TipToe.TheWhiteWest)
        for col in ["A4", "E4"]:
            InsightsWS[col] = ""
            TipToe.Flex(InsightsWS[col], **TipToe.TheWhiteWest)

        Features = [
            "BOX TEMP (Celsius):", "RADIOACTIVE DECAY RATE:", "PHOTON COUNT (per min):",
            "WAVEFUNCTION STABILITY:", "ENTANGLEMENT INDEX:", "OBSERVER PRESENCE:",
            "CARDBOARD BOX:", "LEAD BOX:", "QUANTUM FOAM BOX:", "VELVET BOX:",
        ]
        for i, label in enumerate(Features, start=5): #FILL IN THE NAMES OF THE FEATURES A5:A14
            InsightsWS[f"A{i}"] = label 
            TipToe.Flex(InsightsWS[f"A{i}"], **TipToe.CoolBlueJewels) #STYLE FEATURES IN A5:A14

        InsightsWS.merge_cells("A15:E15") #SPACER
        InsightsWS["A16"] = "MODEL INSIGHTS" #SECTION WITH METRICS ABOUT MODEL'S PERFORMANCE
        TipToe.Flex(InsightsWS["A16"], **TipToe.PurpleIconChopped) #STYLE MODEL INSIGHTS SECTION TITLE
        InsightsWS.merge_cells("A16:E16")

        MetricLabels = [
            "Coefficient of determination:",
            "Intercept:",
            "Mean Absolute Error:",
            "Mean Squared Error:",
            "Root Mean Squared Error:",
            "Accuracy:",
            "Precsion:",
            "Recall:",
            "F1:",
            "AUC:"
        ]
        for i, metric in enumerate(MetricLabels, start=18): #FILL IN THE NAMES OF THE METRICS IN A18:A27
            InsightsWS[f"A{i}"] = metric
            TipToe.Flex(InsightsWS[f"A{i}"], **TipToe.CoolBlueJewels) #STYLE THE METRICS

        MetricHeaders = [ 
            ("A17", "METRIC"),
            ("B17", "MOOD"),
            ("C17", "SASS"),
            ("D17", "SURVIVAL"),
            ("E17", "INSIGHTS")
        ]
        for col, text in MetricHeaders:
            InsightsWS[col] = text
            TipToe.Flex(InsightsWS[col], **TipToe.TheWhiteWest) #STYLE THE METRICS IN B-D

        Drip = TipToe(self.wb, InsightsWS) #APPLY DRIP
        Drip.MetricsDrip()
        Drip.InsightsDrip()
       

    def GetCoefficientsAndMetrics(self, CoefficientsDF=None, MetricsDF=None): #GRAB THE FEATURE IMPORTANCES AND THE MODEL PERFORMANCE METRICS
        ws = self.GetWS("INSIGHTS")
        if CoefficientsDF is not None:
            for i, row in CoefficientsDF.iterrows():
                ws[f"B{i+5}"] = row["MoodImportance"]
                ws[f"C{i+5}"] = row["SassImportance"]
                ws[f"D{i+5}"] = row["SurvivalImportance"]
        if MetricsDF is not None:
            for i, row in MetricsDF.iterrows():
                rownum = 18 + i
                ws[f"B{rownum}"] = row["Mood"]
                ws[f"C{rownum}"] = row["Sass"]
                ws[f"D{rownum}"] = row["Survival"]
                ws[f"E{rownum}"] = row["ModelInsights"]
        self.SAVE()

    def CatWisdom(self, CoefficientsDF, MetricsDF): #USE THE COEFFICIENTS AND METRICS TO GENERATE INSIGHTS ABOUT THE CATS AND THE MODEL
        MoodFeatures = ["BoxTemp", "DecayRate", "Photons", "Stability",
                        "Material_Cardboard", "Material_Lead", "Material_Velvet", "Material_QuantumFoam"]
        SassFeatures = ["BoxTemp", "Entanglement"]
        SurvivalFeatures = ["Observer"] #FEATURE GROUPINGS

        for feature in MoodFeatures: #MOOD
            MoodCoefficient = float(self.Importances.get("Mood", {}).get(feature, 0.0))
            if feature == "BoxTemp":
                if MoodCoefficient >= 0.6:
                    MoodInsight = "Warm boxes put cats in good moods 😻🔥"
                elif MoodCoefficient >= -0.5:
                    MoodInsight = "Cats moods are not affected by temperature 😐"
                elif MoodCoefficient >= -1.9:
                    MoodInsight = "Cold boxes put cats in bad moods 🙀❄️"


            elif feature == "DecayRate":
                if MoodCoefficient >= 0.6:
                    MoodInsight = "Cats LOVE risky situations 😹☢️"
                elif MoodCoefficient >= -0.5:
                    MoodInsight = "Cats are neutral about radioactive chaos 😐"
                else:
                    MoodInsight = "Cats like to play it safe 🙀⚛️"

            elif feature == "Photons":
                if MoodCoefficient >= 0.6:
                    MoodInsight = "Cats LOVE fast-paced quantum activity 😹⚡"
                elif MoodCoefficient >= -0.5:
                    MoodInsight = "Cats ignore the flickering photons 😐"
                else:
                    MoodInsight = "Cats prefer a chill quantum vibe 😸💤"

            elif feature == "Stability":
                if MoodCoefficient >= 0.6:
                    MoodInsight = "Cats LOVE feeling stable 😺✨"
                elif MoodCoefficient >= -0.5:
                    MoodInsight = "Cats are indifferent to their stability 😐"
                else:
                    MoodInsight = "Cats don’t like instability 🙀⚛️"

            else:  # BOX MATERIALS
                if MoodCoefficient >= 2.0:
                    MoodInsight = f"Cats LOVE {feature.replace('Material_', '')} boxes 😻"
                elif MoodCoefficient >= 0.6:
                    MoodInsight = f"Cats like {feature.replace('Material_', '')} boxes 😺"
                elif MoodCoefficient >= -0.5:
                    MoodInsight = f"Cats are indifferent to {feature.replace('Material_', '')} boxes 😐"
                elif MoodCoefficient >= -1.9:
                    MoodInsight = f"Cats dislike {feature.replace('Material_', '')} boxes 🙀"
                else:
                    MoodInsight = f"Cats HATE {feature.replace('Material_', '')} boxes 😾"

            CoefficientsDF.loc[CoefficientsDF["Feature"] == feature, "MoodImportance"] = MoodCoefficient
            CoefficientsDF.loc[CoefficientsDF["Feature"] == feature, "FeatureInsights"] += MoodInsight

        for feature in SassFeatures: #SASS
            SassCoefficient = float(self.Importances.get("Sass", {}).get(feature, 0.0))
            if feature == "BoxTemp":
                if SassCoefficient >= 0.6:
                    SassInsight = "Cats get sassy in higher box temps 😺🔥"
                elif SassCoefficient >= -0.5:
                    SassInsight = "Temperature doesn’t affect sass 😐"
                else:
                    SassInsight = "Too cold to sass 🙀❄️"
            elif feature == "Entanglement":
                if SassCoefficient >= 0.6:
                    SassInsight = "External systems make cats extra sassy 😹⚛️"
                elif SassCoefficient >= -0.5:
                    SassInsight = "Cats are neutral about entanglement 😐"
                else:
                    SassInsight = "Cats don’t care about quantum links 😾"

            CoefficientsDF.loc[CoefficientsDF["Feature"] == feature, "SassImportance"] = SassCoefficient
            ExistingInsight = CoefficientsDF.loc[CoefficientsDF["Feature"] == feature, "FeatureInsights"].values[0] #ADD TO EXISTING INSIGHT FOR MOOD / BOX TEMP
            AdditionalInsight = f"{ExistingInsight} | {SassInsight}" if ExistingInsight else SassInsight
            CoefficientsDF.loc[CoefficientsDF["Feature"] == feature, "FeatureInsights"] = AdditionalInsight

        for feature in SurvivalFeatures: #SURVIVAL
            SurvivalCoefficient = float(self.Importances.get("Survival", {}).get(feature, 0.0))
            if SurvivalCoefficient >= 0.6:
                SurvivalInsight = "Cats LOVE the head pats 😻🙌"
            elif SurvivalCoefficient >= -0.5:
                SurvivalInsight = "Cats allow humans to exist 😺"
            else:
                SurvivalInsight = "Just leave them alone 😾📦"

            CoefficientsDF.loc[CoefficientsDF["Feature"] == feature, "SurvivalImportance"] = SurvivalCoefficient
            CoefficientsDF.loc[CoefficientsDF["Feature"] == feature, "FeatureInsights"] += f"{SurvivalInsight}"

        try: #WRITE TO EXCEL
            ws = self.wb["INSIGHTS"]  # INSIGHTS TAB
        except KeyError:
            ws = self.wb.create_sheet("INSIGHTS")

        start_row = 5
        for i, text in enumerate(CoefficientsDF["FeatureInsights"], start=start_row):
            ws[f"E{i}"] = text


        start_row = 18
        for i, text in enumerate(MetricsDF["ModelInsights"], start=start_row):
            ws[f"E{i}"] = text
      

        Drip = TipToe(self.wb, ws)
        Drip.MetricsDripChopped()
        self.SAVE()

    def DataKitten(self, ws=None): #FILLS IN DATA TAB WITH THE INPUT DF AND STYLE IT
        ws = self.GetWS("DATA")
        DataTabHeaders = ["BOX\nTEMPERATURE\n(Celsius)", "RADIOACTIVE\nDECAY RATE", "PHOTON COUNT\n(PER MINUTE)", "WAVEFUNCTION\nSTABILITY", "ENTANGLEMENT\nINDEX", "OBSERVER\nPRESENT?", "BOX\nMATERIAL", "ACTUAL\nMOOD SCORE", "PREDICTED\nMOOD SCORE", "MOOD SCORE\nRESIDUAL", "ACTUAL\nSASS INDEX", "PREDICTED\nSASS INDEX", "SASS INDEX\nRESIDUAL", "ACTUAL\nSURVIVAL", "PREDICTED\nSURVIVAL", "SURVIVAL RATE\nRESIDUAL"] #Data tab headers in order
        for i, header in enumerate(DataTabHeaders, start=1): #PRETTY HEADERS
            cell = ws.cell(row=1, column=i, value=header)
        
        for r, row in enumerate(self.df.itertuples(index=False), start=2): #PASTE DF INTO EXCEL / DATA TAB
            for c, value in enumerate(row, start=1):
                ColumnName = self.df.columns[c - 1] #OBSERVER COLUMN (YES / NO)
                if ColumnName == "Observer":
                    GlamObserver = str(value) if pd.notna(value) else ""
                else:
                    GlamObserver = value
                ws.cell(row=r, column=c, value=GlamObserver)
        Drip = TipToe(self.wb, ws) 
        Drip.DataDrip()
        self.SAVE()

    def PlaceImage(self, ws, fig, AnchorCell, HeightIN, WidthIN): #SAFELY SAVES MATPLOTLIB AS TEMP PNG, INSERTS INTO APPROPRIATE TAB, HANDLES MISSING IMAGE ERRORS GRACEFULLY, DEFERS DELETION OF TEMPS UNTIL AFTER SAVE
        global atexit
        TempPath = None
        try: #CREATE A  TEMP FILE BUT DON"T AUTO-DELETE - WINDOWS LOCKS
            with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmpfile:
                TempPath = tmpfile.name
            fig.savefig(TempPath, format="png", dpi=96, bbox_inches="tight")  #SAVE FIGURE
            #print(f"[DEBUG] Inserted {TempPath}")
            if os.path.exists(TempPath) and os.path.getsize(TempPath) > 0: #VERIFY AND INSERT
                img = Image(TempPath)
                img.width = int(WidthIN * 96)
                img.height = int(HeightIN * 96)
                img.anchor = AnchorCell
                ws.add_image(img)
            else:
                ws[AnchorCell] = "Image Not Found"
            if TempPath: #CLEAN UP TEMPS AFTER PROGRAM EXITS
                atexit.register(lambda path=TempPath: os.path.exists(path) and os.remove(path))
        except Exception as e:
            ws[AnchorCell] = f"Image Not Found ({e.__class__.__name__})"
            #print(f"[DEBUG] Inserted {TempPath}")

    def PlaceScatter(self, ws, fig, AnchorCell, HeightIN, WidthIN): #PLACE SCATTER PLOTS
        self.PlaceImage(ws, fig, AnchorCell, HeightIN, WidthIN)

    def ScatterKitten(self, ws=None): #FOR SCATTER PLOTS TAB
        ScatterWS = self.GetWS("SCATTER PLOTS")
        ScatterWS["A1"] = "SCATTER PLOTS"
        ScatterWS.merge_cells("A1:G1")
        ScatterWS["B2"] = "MOOD SCORE"
        ScatterWS["D2"] = "SASS INDEX"
        ScatterWS["F2"] = "SURVIVAL RATE"
        # --- MOOD ---
        self.PlaceScatter(ScatterWS, self.ScatterPlots["Box Temperature vs Mood Score"], "B4", 2.18, 3.83)
        self.PlaceScatter(ScatterWS, self.ScatterPlots["Decay Rate vs Mood Score"], "B16", 2.18, 3.83)
        self.PlaceScatter(ScatterWS, self.ScatterPlots["Photon Count vs Mood Score"], "B28", 2.18, 3.83)
        self.PlaceScatter(ScatterWS, self.ScatterPlots["Stability vs Mood Score"], "B40", 2.18, 3.83)
        self.PlaceScatter(ScatterWS, self.ScatterPlots["Material vs Mood Score"], "B52", 2.18, 3.83)
        # --- SASS ---
        self.PlaceScatter(ScatterWS, self.ScatterPlots["Box Temperature vs Sass Index"], "D4", 2.18, 3.83)
        self.PlaceScatter(ScatterWS, self.ScatterPlots["Entanglement vs Sass Index"], "D16", 2.18, 3.83)
        # --- SURVIVAL ---
        self.PlaceScatter(ScatterWS, self.ScatterPlots["Observer Presence vs Survival Rate"], "F4", 2.18, 3.83)
        
        Drip = TipToe(self.wb, ScatterWS)
        Drip.PlotsDrip()
        
    def MoodResultsKitten(self, ws=None): #FOR LINEAR REGRESSION PICS FOR MOOD SCORE
        MoodWS = self.wb["MOOD RESULTS"]
        MoodWS["A1"] = "LINEAR REGRESSION"
        MoodWS.merge_cells("A1:I1")
        MoodWS["A2"] = "MOOD SCORE"
        MoodWS.merge_cells("A2:I2")

        key = "Box Temperature vs Mood Score"
        if key in self.RegressionPlots and self.RegressionPlots[key] is not None:
            self.PlaceImage(MoodWS, self.RegressionPlots["Box Temperature vs Mood Score"], "B4", 3.02, 6.1)
        else:
            MoodWS["B4"] = "No Mood / Temp  model available"

        key = "Decay Rate vs Mood Score"
        if key in self.RegressionPlots and self.RegressionPlots[key] is not None:
            self.PlaceImage(MoodWS, self.RegressionPlots["Decay Rate vs Mood Score"], "F4", 3.02, 6.1)
        else:
            MoodWS["F4"] = "No Mood / Decay model available"

        key = "Photon Count vs Mood Score"
        if key in self.RegressionPlots and self.RegressionPlots[key] is not None:
            self.PlaceImage(MoodWS, self.RegressionPlots["Photon Count vs Mood Score"], "B20", 3.02, 6.1)
        else:
            MoodWS["B20"] = "No Mood / Photon  model available"

        key = "Stability vs Mood Score"
        if key in self.RegressionPlots and self.RegressionPlots[key] is not None:
            self.PlaceImage(MoodWS, self.RegressionPlots["Stability vs Mood Score"], "F20", 3.02, 6.1)
        else:
            MoodWS["F20"] = "No Mood / Stability model available"

        key = "Cardboard Box vs Mood Score"
        if key in self.RegressionPlots and self.RegressionPlots[key] is not None:
            self.PlaceImage(MoodWS, self.RegressionPlots["Cardboard Box vs Mood Score"], "B36", 2.2, 2.95)
        else:
            MoodWS["B36"] = "No Mood / Cardboard model available"

        key = "Lead Box vs Mood Score"
        if key in self.RegressionPlots and self.RegressionPlots[key] is not None:
            self.PlaceImage(MoodWS, self.RegressionPlots["Lead Box vs Mood Score"], "D36", 2.2, 2.95)
        else:
            MoodWS["D36"] = "No Mood / Lead model available"

        key = "Quantum Foam Box vs Mood Score"
        if key in self.RegressionPlots and self.RegressionPlots[key] is not None:
            self.PlaceImage(MoodWS, self.RegressionPlots["Quantum Foam Box vs Mood Score"], "F36", 2.2, 2.95)
        else:
            MoodWS["F36"] = "No Mood / Foam model available"

        key = "Velvet Box vs Mood Score"
        if key in self.RegressionPlots and self.RegressionPlots[key] is not None:
            self.PlaceImage(MoodWS, self.RegressionPlots["Velvet Box vs Mood Score"], "H36", 2.2, 2.95)
        else:
            MoodWS["H36"] = "No Mood / Velvet model available"

        Drip = TipToe(self.wb, MoodWS)
        Drip.PlotsDrip()
        
    def SassResultsKitten(self, ws=None): #FOR LINEAR REGRESSION PICS FOR SASS INDEX
        SassWS = self.wb["SASS RESULTS"]
        SassWS["A1"] = "LINEAR REGRESSION PLOTS"
        SassWS.merge_cells("A1:E1")
        SassWS["A2"] = "SASS INDEX"
        SassWS.merge_cells("A2:E2")

        key = "Box Temperature vs Sass Index"
        if key in self.RegressionPlots and self.RegressionPlots[key] is not None:
            self.PlaceImage(SassWS, self.RegressionPlots["Box Temperature vs Sass Index"], "B4", 3.22, 5.78)
        else:
            SassWS["B4"] = "No Sass / Box Temp model available"

        key = "Entanglement vs Sass Index"
        if key in self.RegressionPlots and self.RegressionPlots[key] is not None:
            self.PlaceImage(SassWS, self.RegressionPlots["Entanglement vs Sass Index"], "D4", 3.22, 5.78)
        else:
            SassWS["D4"] = "No Sass / Entanglement  model available"

        Drip = TipToe(self.wb, SassWS)
        Drip.PlotsDrip()
        self.SAVE()

    def SurvivalResultsKitten(self, ws=None): #FOR LINEAR REGRESSION PICS FOR SURVIVAL RATE
        SurvivialWS = self.wb["SURVIVAL RESULTS"]
        SurvivialWS["A1"] = "LINEAR REGRESSION PLOTS"
        SurvivialWS.merge_cells("A1:C1")
        SurvivialWS["A2"] = "SURVIVAL RATE"
        SurvivialWS.merge_cells("A2:C2")

        key = "Observer Presence vs Survival Rate"
        if key in self.RegressionPlots and self.RegressionPlots[key] is not None:
            self.PlaceImage(SurvivialWS, self.RegressionPlots[key], "B4", 3.63, 6.88)
        else:
            SurvivialWS["B4"] = "No survival model available"

        Drip = TipToe(self.wb, SurvivialWS)
        Drip.PlotsDrip()
        
    def ExcelLitter(self, OutputPath):
        self.InsightsKitten() 
        #self.CatWisdom()
        #self.DataKitten() 
        self.ScatterKitten() 
        self.MoodResultsKitten()
        self.SassResultsKitten()
        self.SurvivalResultsKitten()
        self.SAVE()


