import pandas as pd
import sys
import os
import io
import tkinter as tk
import subprocess
import re
import tempfile
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image as XLImage
import numpy as np
from io import BytesIO
from tkinter import Tk, Label, Entry, Button, filedialog, simpledialog, messagebox, ttk
from tkinter import *
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.drawing.image import Image
from CatDrip import DrippyKit
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import OneHotEncoder
from sklearn.linear_model import LinearRegression
from sklearn.metrics import r2_score, mean_squared_error
import seaborn as sns

#---------------------------- CLASSES TO COLLECT INFO, CLEAN DATA, AND GET THE HEADERS ---------------------------- 
DATA_DF_HEADERS = [
    "BoxTemp",
    "Photons",
    "Entanglement",
    "Observer",
    "DecayRate",
    "Stability",
    "Material"
]

INPUT_HEADERS = {
    "Box Temperature (Â°C)|Box Temperature (°C)|Box_Temp|box temp": "BoxTemp",
    "Photon Count per Minute": "Photons",
    "Quantum Entanglement Index": "Entanglement", 
    "Observer Presence": "Observer", 
    "Radioactive Decay Rate": "DecayRate", 
    "Wavefunction Stability": "Stability",
    "Box Material": "Material"
    }

DATA_TAB_HEADERS = {
    "BoxTemp":"BOX TEMP\n(Celsius)", 
    "Photons": "PHOTON COUNT\n(per  minute)", 
    "Entanglement": "ENTANGLEMENT\nINDEX", 
    "Observer": "OBSERVER\nPRESENT?", 
    "DecayRate": "RADIOACTIVE\nDECAY RATE", 
    "Stability": "WAVEFUNCTION\nSTABILITY", 
    "Material": "BOX\nMATERIAL"
    }

RegresionMap = {
    "BoxTemp|BOX TEMP\n(Celsius)|Box Temperature (Â°C)|Box Temperature (°C)": "BoxTemp",
    "Photons|Photon Count per Minute|PHOTON COUNT\n(per  minute)": "Photons",
    "Entanglement|Quantum Entanglement Index|ENTANGLEMENT\nINDEX": "Entanglement",
    "Observer|Observer Presence|OBSERVER\nPRESENT?": "Observer",
    "DecayRate|Radioactive Decay Rate|RADIOACTIVE\nDECAY RATE": "DecayRate",
    "Stability|Wavefunction Stability|WAVEFUNCTION\nSTABILITY": "Stability",
    "Material|Box Material|BOX\nMATERIAL": "Material",
    "MoodScore": "MoodScore",
    "SassIndex": "SassIndex",
    "SurvivalRate": "SurvivalRate"
    }

REGRESSION_HEADERS = {}
for variants, canonical in RegresionMap.items():
    for variant in variants.split("|"):
        REGRESSION_HEADERS[variant.strip()] = canonical

REGRESSION_TAB_HEADERS = {
    "BoxTemp":"BOX TEMP\n(Celsius)", 
    "Photons": "PHOTON COUNT\n(per  minute)", 
    "Entanglement": "ENTANGLEMENT\nINDEX", 
    "Observer": "OBSERVER\nPRESENT?", 
    "DecayRate": "RADIOACTIVE\nDECAY RATE", 
    "Stability": "WAVEFUNCTION\nSTABILITY", 
    "Material": "BOX\nMATERIAL",
    "MoodScore": "MOOD\nSCORE",
    "SassIndex": "SASS\nINDEX",
    "SurvivalRate": "SURVIVAL\nRATE"
    }

ImportanceRenameMap = {
	"BoxTemp":"BOX TEMP (Celsius)", 
   	"Photons": "PHOTON COUNT (per  minute)", 
	"Entanglement": "ENTANGLEMENT INDEX", 
	"Observer": "OBSERVER PRESENCE", 
	"DecayRate": "RADIOACTIVE DECAY RATE", 
	"Stability": "WAVEFUNCTION STABILITY", 
	"Material_Graphene": "GRAPHENE BOX",
	"Material_Lead": "LEAD BOX",
	"Material Quantum Foam": "QUANTUM FOAM BOX",
	"Material_Velvet": "VELVET BOX"
	}

TARGET_HEADERS = ["MoodScore", "SassIndex", "SurvivalRate"]


#---------------------------- DATA/INFO CLASSES ---------------------------- 

class PurrfectData: #Clean Data
    def __init__(self, df, InputHeaders, DFHeaders, OutputHeaders):
        self.df = df.copy()
        self.InputHeaders = InputHeaders
        self.DFHeaders = DFHeaders
        self.OutputHeaders = OutputHeaders
    
    def CleanDataRightMeow(self):
        self.df = self.df.drop_duplicates()

        numeric_cols = self.df.columns[:6]  # Columns A–F
        string_col = self.df.columns[6]     # Column G

        self.df[numeric_cols] = self.df[numeric_cols].apply(pd.to_numeric, errors='coerce')
        self.df[string_col] = self.df[string_col].astype(str)

        self.df = self.df.dropna() #Handle missing values (none expected, but good practice) or use 

        #valid_materials = ['Cardboard', 'Lead', 'Graphene', 'Velvet', 'Quantum Foam'] #Validate categorical values
        #df = df[df['box_material'].isin(valid_materials)]

        #df['observer_presence'] = df['observer_presence'].astype(bool) #Convert observer presence to boolean
    

    def normalize_headers(self):
        rename_map = {}
        for aliases, standard_name in self.InputHeaders.items():
            for alias in aliases.split("|"):
                if alias in self.df.columns:
                    rename_map[alias] = standard_name
        self.df = self.df.rename(columns=rename_map)
        print("Raw columns before normalization:", self.df.columns.tolist())

    def groom(self):
        self.CleanDataRightMeow()
        self.normalize_headers()
        self.df = self.df.rename(columns=DATA_TAB_HEADERS)
        return self.df

class GetInfo:
   
    def GetUserInfo():
        UserInfo = {}

        def on_submit():
            UserInfo["Name"] = Name.get()
            UserInfo["DateToday"] = datetime.today().strftime("%Y%m%d")
            root.quit()
            root.destroy()

        def on_cancel():
            UserInfo["cancelled"] = True
            root.quit()
            root.destroy()
            print("User Canceled")
            sys.exit(1)

        root = Tk()
        root.title("Who are you?")
        root.geometry("300x100")
        root.configure(bg="#C2CAE8")

        label_style = {"bg": "#C2CAE8", "fg": "#000000", "font": ("Courier New", 11, "bold")}
        entry_style = {"bg": "#ffffff", "fg": "#000000", "font": ("Arial", 11)}

        # Name label and entry
        Label(root, text="Your Name:", **label_style).grid(row=0, column=0, sticky="e", padx=10, pady=10)
        Name = Entry(root, width=20, **entry_style)
        Name.grid(row=0, column=1, sticky="w", padx=10, pady=10)

        # Button frame for centering
        button_frame = Frame(root, bg="#C2CAE8")
        button_frame.grid(row=1, column=0, columnspan=2, pady=10)

        Button(
            button_frame,
            text="Cancel",
            command=on_cancel,
            bg="#CBCEDF",
            fg="#000000",
            font=("Arial", 11),
            relief="raised",
            padx=15, pady=5
        ).pack(side=LEFT, padx=10)

        Button(
            button_frame,
            text="Submit",
            command=on_submit,
            bg="#CBCEDF",
            fg="#000000",
            font=("Arial", 11),
            relief="raised",
            padx=15, pady=5
        ).pack(side=RIGHT, padx=10)

        root.eval('tk::PlaceWindow . center')
        root.mainloop()

        return UserInfo

    def GetInputFile():
        if len(sys.argv) > 1:
            # Drag-and-drop case
            InputFile = sys.argv[1]
            if not os.path.isfile(InputFile):
                print("The dropped file is not valid.")
                sys.exit(1)
        else:
            # No drag-and-drop → open file picker
            root = Tk()
            root.withdraw()  # hide main Tkinter window
            InputFile = filedialog.askopenfilename(
                title="Select CSV file",
                filetypes=[("CSV Files", "*.csv")]
            )
            if not InputFile:
                print("No file selected. Exiting.")
                sys.exit(0)

        print(f"Processing file: {InputFile}")
        return InputFile

    def GetDF(InputFile):
        df = pd.read_csv(InputFile, encoding='utf-8')
        return df

    def GetSavePath(UserInput):
        """Generate and verify save path using subscription reference ID from B3."""
    # Load subscription reference ID from Summary tab
        
        UserName = UserInput["Name"]
        BaseName = f"Cat_Cafe_Result_{UserName}_{UserInput['DateToday']}.xlsx"

        Tk().withdraw()
        SavePath = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=BaseName)

        if SavePath and os.path.exists(SavePath):
            version = simpledialog.askstring("Version", "Enter version number:") or "2"
            BaseName = f"Cat_Cafe_Result_{UserName}_{UserInput['DateToday']}_v{version}.xlsx"
            SavePath = os.path.join(os.path.dirname(SavePath), BaseName)

        OutputFileName = SavePath
        return OutputFileName

#---------------------------- REGRESSION CLASS ---------------------------- 
class QuantumCafeModel:
    def __init__(self, df):
        self.df = df.copy()
        self.model = None
        self.feature_columns = ['BoxTemp', 'Photons', 'Entanglement',
                                'Observer', 'DecayRate', 'Stability', 'Material']

    def encode_features(self):
        """One-hot encode categorical Material column."""
        X = self.df[self.feature_columns]
        X = pd.get_dummies(X, columns=['Material'], drop_first=True)
        return X

    def split(self, target_col, test_size=0.2, random_state=42):
        """Split data into train and test sets."""
        X = self.encode_features()
        y = self.df[target_col]
        return train_test_split(X, y, test_size=test_size, random_state=random_state)

    def train(self, target_col, model=None):
        """Train a regression model on the given target."""
        X_train, X_test, y_train, y_test = self.split(target_col)

        # Use given model or default to LinearRegression
        self.model = model if model else LinearRegression()
        self.model.fit(X_train, y_train)

        # Predictions and metrics
        y_pred = self.model.predict(X_test)
        r2 = r2_score(y_test, y_pred)
        mse = mean_squared_error(y_test, y_pred)

        print(f"📊 Results for {target_col}:")
        print(f"R²: {r2:.3f}")
        print(f"MSE: {mse:.3f}")

        return self.model, (r2, mse)

    def feature_importance(self, target_col):
        """Print feature coefficients/importance (for linear models)."""
        X = self.encode_features()
        feature_names = X.columns

        if hasattr(self.model, "coef_"):
            coefs = pd.Series(self.model.coef_, index=feature_names)
            print(f"\n🔍 Feature importance for {target_col}:")
            print(coefs.sort_values(ascending=False))
        else:
            print("This model does not provide coefficients.")

#---------------------------- GET INPUT AND DATA ----------------------------
#UserInfo = GetInfo.GetUserInfo()
#InputFile = GetInfo.GetInputFile()
InputFile = "cat_cafe_dataset.csv"
Purr = PurrfectData(
    df = GetInfo.GetDF(InputFile),
    DFHeaders=DATA_DF_HEADERS,
    InputHeaders=INPUT_HEADERS,
    OutputHeaders=DATA_TAB_HEADERS
)
DataDF = Purr.groom()

#---------------------------- REGRESSION DATA FRAME ---------------------------- 
UpdateDataDF = DataDF
UpdateDataDF["MoodScore"] = np.random.uniform(0, 1, len(UpdateDataDF))
UpdateDataDF["SassIndex"] = np.random.randint(0, 10, len(UpdateDataDF))
UpdateDataDF["SurvivalRate"] = np.random.normal(loc=0.8, scale=0.1, size=len(UpdateDataDF)).clip(0, 1)


missing = [
    canonical for canonical, aliases in INPUT_HEADERS.items()
    if not any(alias in UpdateDataDF.columns for alias in aliases.split('|'))
]
print("Missing columns:", missing)

missing_targets = [target for target in TARGET_HEADERS if target not in UpdateDataDF.columns]
if missing_targets:
    print("Missing target columns:", missing_targets)



RegressionDF = UpdateDataDF.rename(columns=REGRESSION_HEADERS)
print("Final columns in RegressionDF:", RegressionDF.columns.tolist())

CatWhisperer = QuantumCafeModel(RegressionDF) # Initialize with dataframe
CatWhisperer.train("MoodScore") # Train for MoodScore
CatWhisperer.train("SurvivalRate") # Train for SurvivalRate
CatWhisperer.train("SassIndex") # Train for SassIndex
CatWhisperer.feature_importance("MoodScore") # Look at feature influence for MoodScore

#OutputFileName = GetInfo.GetSavePath(UserInfo)
OutputFileName = "output.xlsx"
if not OutputFileName:
    print("No save location chosen. Exiting.")

#df.to_excel('OutputFileName.xlsx', index=False)
with pd.ExcelWriter(OutputFileName, engine="openpyxl") as writer:
    RegressionDF.to_excel(writer, sheet_name="CAT_REGRESSION", index=False)
    #DataDF.to_excel(writer, sheet_name="CAT_DATA", index=False)
    
DataDrip = DrippyKit(filepath=OutputFileName, sheet_name="CAT_REGRESSION")
HeaderRow = next(DataDrip.ws.iter_rows(min_row=1, max_row=1))
LastColumn = DataDrip.ws.max_column + 1
FirstItemsRow = 2  # Just use the row number
LastRow = DataDrip.ws.max_row
DataDrip.HeaderLewk()
DataDrip.ItemsLewk()
DataDrip.ColumnWidths()
DataDrip.ThiccBorder()
DataDrip.wb.save(OutputFileName)

#drip = DrippyKit(filepath=OutputFileName, sheet_name="CAT_DATA")
#HeaderRow = next(drip.ws.iter_rows(min_row=1, max_row=1))
#LastColumn = drip.ws.max_column + 1
#FirstItemsRow = 2  # Just use the row number
#LastRow = drip.ws.max_row
#drip.HeaderLewk()
#drip.ItemsLewk()
#drip.ColumnWidths()
#drip.ThiccBorder()
#drip.wb.save(OutputFileName)

#---------------------------- SAVE ----------------------------
wb = load_workbook(OutputFileName)
wb.save(OutputFileName)
print(f"DONE")
#print("DataFrame Headers:", list(DataDF.columns))
wb = load_workbook(OutputFileName) # Load Excel back in with openpyxl
wsREGRESSION = wb["CAT_REGRESSION"]
excel_regression_headers = [cell.value for cell in wsREGRESSION[1]]
#wsDATA = wb["CAT_DATA"]
#excel_data_headers = [cell.value for cell in wsDATA[1]]  # First row is headers
#----------------------------  ADD SHEETS ---------------------------- 
if "SCATTER_PLOTS" not in OutputFileName:
    wsScatterPlots = wb.create_sheet("SCATTER_PLOTS")
else:
    wsScatterPlots = wb["SCATTER_PLOTS"]

#----------------------------  MODELING - SCATTER PLOTS ---------------------------- 
wsScatterPlots.column_dimensions["A"].width = 50
wsScatterPlots.column_dimensions["B"].width = 3
wsScatterPlots.column_dimensions["C"].width = 50
wsScatterPlots.column_dimensions["D"].width = 3
wsScatterPlots.column_dimensions["E"].width = 50

QuantumFeatures = ["BoxTemp", "Photons", "Entanglement", "Observer", "DecayRate", "Stability", "Material"]
QuantumTargets = ["MoodScore", "SassIndex", "SurvivalRate"]

TargetFeatureMap = {
    "MoodScore": ["BoxTemp", "Photons", "Entanglement", "DecayRate", "Stability", "Material"],
    "SassIndex": ["BoxTemp", "Entanglement"],
    "SurvivalRate": ["Observer"]
}

row = 1
temp_files = []
# Column mapping for each target
col_map = {
    "MoodScore": "A",      # First strip
    "SassIndex": "C",      # Skip one column
    "SurvivalRate": "E"    # Skip again
}

row_map = {"MoodScore": 1, "SassIndex": 1, "SurvivalRate": 1}  # Track row per target

for target in QuantumTargets:
    for feature in QuantumFeatures:
        if feature in RegressionDF.columns:
            if feature == "Material":
                fig, ax = plt.subplots(figsize=(3.75, 3.5))
            else:
                fig, ax = plt.subplots(figsize=(3.75, 2.5))
            ax.scatter(RegressionDF[feature], RegressionDF[target], alpha=0.5, color="#4B1395", s=15, label="Data")
            ax.set_title(f"{feature} vs {target}")
            ax.set_xlabel(feature)
            ax.set_ylabel(target)
            if feature == "Material":
                plt.setp(ax.get_xticklabels(), rotation=45, ha="right", fontsize=8)
            ax.grid(True)
            plt.tight_layout()

            # Save image to temp file
            tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
            tmp_path = tmp.name
            tmp.close()
            plt.savefig(tmp_path)
            plt.close(fig)
            temp_files.append(tmp_path)

            # Place image in Excel at the correct column/row
            img = XLImage(tmp_path)
            col = col_map[target]
            row = row_map[target]
            wsScatterPlots.add_image(img, f"{col}{row}")

            # Move down for the next feature plot of this target
            row_map[target] += 13

wb.save(OutputFileName) # Save workbook before cleanup

for path in temp_files: #  delete temp files
    os.remove(path)

#  ---------------------------- REGRESSION HELPER FUNCTION ----------------------------
def plot_regression(ws, RegressionDF, feature, target, start_row, col="A"):
    """Create scatter + regression line plot for one feature vs. target, insert into Excel."""
    fig, ax = plt.subplots(figsize=(6,4))

    # Scatter plot
    ax.scatter(
        RegressionDF[feature], RegressionDF[target],
        alpha=0.6, color="#4B1395", s=15, label="Data"
    )

    # Fit regression line
    X_feat = RegressionDF[[feature]]  # keep as DataFrame (preserves column name)
    lr = LinearRegression()
    lr.fit(X_feat, RegressionDF[target])

    x_vals = np.linspace(X_feat.min().values[0], X_feat.max().values[0], 100)
    x_df = pd.DataFrame(x_vals, columns=[feature])  # keep column name!
    y_vals = lr.predict(x_df)

    ax.plot(x_vals, y_vals, color="red", linewidth=2, label="Regression Line")

    # Labels + style
    ax.set_title(f"{feature} vs {target}")
    ax.set_xlabel(feature)
    ax.set_ylabel(target)
    ax.grid(True)
    ax.legend()
    plt.tight_layout()

    # Save temporary file
    tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
    plt.savefig(tmp.name)
    plt.close(fig)

    # Insert image into Excel
    img = XLImage(tmp.name)
    ws.add_image(img, f"{col}{start_row}")

    return tmp.name  # return path for cleanup

#  ---------------------------- MODELING - MOOD SCORE REGRESSION ----------------------------
wb = load_workbook(OutputFileName)

if "MOOD_REGRESSION" not in wb.sheetnames:
    wsMood = wb.create_sheet("MOOD_REGRESSION")
else:
    wsMood = wb["MOOD_REGRESSION"]

wsMood.column_dimensions["A"].width = 25
wsMood.column_dimensions["B"].width = 25

target = "MoodScore"
features = ["BoxTemp", "Photons", "Entanglement", "DecayRate", "Stability", "Material"]

# Encode only those features (Material expands into dummy columns)
X = RegressionDF[features].copy()
X = pd.get_dummies(X, drop_first=True)
y = RegressionDF[target]

# Fit regression on the encoded features
model = LinearRegression()
model.fit(X, y)

# Extract coefficients with your rename map
coefs = pd.Series(model.coef_, index=X.columns)
ordered = list(ImportanceRenameMap.keys())  # enforce your preferred order
coefs = coefs.reindex(ordered).dropna()

# Headers
wsMood["A1"] = "FEATURE"
wsMood["B1"] = "IMPORTANCE"

# Write coefficients
row = 2
for feat, imp in coefs.items():
    wsMood[f"A{row}"] = ImportanceRenameMap.get(feat, feat)
    wsMood[f"B{row}"] = round(float(imp), 3)
    row += 1

# ---------------- scatter + regression plots ----------------
row = 13
temp_files = []

for feature in features:
    # Encode just the current feature (in case it's categorical)
    X_feat = RegressionDF[[feature]].copy()
    X_feat = pd.get_dummies(X_feat, drop_first=True)

    fig, ax = plt.subplots(figsize=(6, 4))
    ax.scatter(RegressionDF[feature], y, alpha=0.6, color="#4B1395", s=15, label="Data")

    # Fit regression line for this feature
    lr = LinearRegression()
    lr.fit(X_feat, y)

    if X_feat.shape[1] == 1:  # numeric case
        x_vals = np.linspace(X_feat.min().values[0], X_feat.max().values[0], 100)
        x_df = pd.DataFrame(x_vals, columns=X_feat.columns)  # preserve column name(s)
        y_vals = lr.predict(x_df)
        ax.plot(x_vals, y_vals, color="red", linewidth=2, label="Regression Line")

    # Labels
    ax.set_title(f"{feature} vs {target}")
    ax.set_xlabel(feature)
    ax.set_ylabel(target)
    ax.legend()
    ax.grid(True)
    plt.tight_layout()

    # Save temp image
    tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
    tmp_path = tmp.name
    tmp.close()
    plt.savefig(tmp_path)
    plt.close(fig)
    temp_files.append(tmp_path)

    # Insert into Excel
    img = XLImage(tmp_path)
    wsMood.add_image(img, f"A{row}")
    row += 20

wb.save(OutputFileName)

# Clean up temp files
for path in temp_files:
    os.remove(path)




MoodDrip = DrippyKit(filepath=OutputFileName, sheet_name="MOOD_REGRESSION")
MoodDrip.RegressionLewk()


#----------------------------  MODELING - SASS SCORE REGRESSION---------------------------- 
wb = load_workbook(OutputFileName)

if "SASS_REGRESSION" not in wb.sheetnames:
    wsSass = wb.create_sheet("SASS_REGRESSION")
else:
    wsSass = wb["SASS_REGRESSION"]

wsSass.column_dimensions["A"].width = 25
wsSass.column_dimensions["B"].width = 25

target = "SassIndex"
features = ["BoxTemp", "Entanglement"]

# Encode only those features (Material expands into dummy columns)
X = RegressionDF[features].copy()
X = pd.get_dummies(X, drop_first=True)
y = RegressionDF[target]

# Fit regression on the encoded features
model = LinearRegression()
model.fit(X, y)

# Extract coefficients with your rename map
coefs = pd.Series(model.coef_, index=X.columns)
ordered = list(ImportanceRenameMap.keys())  # enforce your preferred order
coefs = coefs.reindex(ordered).dropna()

# Headers
wsSass["A1"] = "FEATURE"
wsSass["B1"] = "IMPORTANCE"

# Write coefficients
row = 2
for feat, imp in coefs.items():
    wsSass[f"A{row}"] = ImportanceRenameMap.get(feat, feat)
    wsSass[f"B{row}"] = round(float(imp), 3)
    row += 1

# ---------------- scatter + regression plots ----------------
row = 13
temp_files = []

for feature in features:
    # Encode just the current feature (in case it's categorical)
    X_feat = RegressionDF[[feature]].copy()
    X_feat = pd.get_dummies(X_feat, drop_first=True)

    fig, ax = plt.subplots(figsize=(6, 4))
    ax.scatter(RegressionDF[feature], y, alpha=0.6, color="#4B1395", s=15, label="Data")

    # Fit regression line for this feature
    lr = LinearRegression()
    lr.fit(X_feat, y)

    if X_feat.shape[1] == 1:  # numeric case
        x_vals = np.linspace(X_feat.min().values[0], X_feat.max().values[0], 100)
        x_df = pd.DataFrame(x_vals, columns=X_feat.columns)  # preserve column name(s)
        y_vals = lr.predict(x_df)
        ax.plot(x_vals, y_vals, color="red", linewidth=2, label="Regression Line")

    # Labels
    ax.set_title(f"{feature} vs {target}")
    ax.set_xlabel(feature)
    ax.set_ylabel(target)
    ax.legend()
    ax.grid(True)
    plt.tight_layout()

    # Save temp image
    tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
    tmp_path = tmp.name
    tmp.close()
    plt.savefig(tmp_path)
    plt.close(fig)
    temp_files.append(tmp_path)

    # Insert into Excel
    img = XLImage(tmp_path)
    wsSass.add_image(img, f"A{row}")
    row += 20

wb.save(OutputFileName)

# Clean up temp files
for path in temp_files:
    os.remove(path)

SassDrip = DrippyKit(filepath=OutputFileName, sheet_name="SASS_REGRESSION")
SassDrip.RegressionLewk()



#----------------------------  MODELING - SURVIVAL RATE REGRESSION---------------------------- 
wb = load_workbook(OutputFileName)

if "SURVIVAL_REGRESSION" not in wb.sheetnames:
    wsAlive = wb.create_sheet("SURVIVAL_REGRESSION")
else:
    wsAlive = wb["SURVIVAL_REGRESSION"]

wsAlive.column_dimensions["A"].width = 25
wsAlive.column_dimensions["B"].width = 25

target = "SurvivalRate"
features = ["Observer"]

# Encode only those features (Material expands into dummy columns)
X = RegressionDF[features].copy()
X = pd.get_dummies(X, drop_first=True)
y = RegressionDF[target]

# Fit regression on the encoded features
model = LinearRegression()
model.fit(X, y)

# Extract coefficients with your rename map
coefs = pd.Series(model.coef_, index=X.columns)
ordered = list(ImportanceRenameMap.keys())  # enforce your preferred order
coefs = coefs.reindex(ordered).dropna()

# Headers
wsAlive["A1"] = "FEATURE"
wsAlive["B1"] = "IMPORTANCE"

# Write coefficients
row = 2
for feat, imp in coefs.items():
    wsAlive[f"A{row}"] = ImportanceRenameMap.get(feat, feat)
    wsAlive[f"B{row}"] = round(float(imp), 3)
    row += 1

# ---------------- scatter + regression plots ----------------
row = 13
temp_files = []

for feature in features:
    # Encode just the current feature (in case it's categorical)
    X_feat = RegressionDF[[feature]].copy()
    X_feat = pd.get_dummies(X_feat, drop_first=True)

    fig, ax = plt.subplots(figsize=(6, 4))
    ax.scatter(RegressionDF[feature], y, alpha=0.6, color="#4B1395", s=15, label="Data")

    # Fit regression line for this feature
    lr = LinearRegression()
    lr.fit(X_feat, y)

    if X_feat.shape[1] == 1:  # numeric case
        x_vals = np.linspace(X_feat.min().values[0], X_feat.max().values[0], 100)
        x_df = pd.DataFrame(x_vals, columns=X_feat.columns)  # preserve column name(s)
        y_vals = lr.predict(x_df)
        ax.plot(x_vals, y_vals, color="red", linewidth=2, label="Regression Line")

    # Labels
    ax.set_title(f"{feature} vs {target}")
    ax.set_xlabel(feature)
    ax.set_ylabel(target)
    ax.legend()
    ax.grid(True)
    plt.tight_layout()

    # Save temp image
    tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
    tmp_path = tmp.name
    tmp.close()
    plt.savefig(tmp_path)
    plt.close(fig)
    temp_files.append(tmp_path)

    # Insert into Excel
    img = XLImage(tmp_path)
    wsAlive.add_image(img, f"A{row}")
    row += 20

wb.save(OutputFileName)

# Clean up temp files
for path in temp_files:
    os.remove(path)



AliveDrip = DrippyKit(filepath=OutputFileName, sheet_name="SURVIVAL_REGRESSION")
AliveDrip.RegressionLewk()


#---------------------------- RENAME REGRESSION HEADERS FOR OUTPUT WORKBOOK ---------------------------- 
wb = load_workbook(OutputFileName)
RegressionDF = RegressionDF.rename(columns=REGRESSION_TAB_HEADERS)
wb.save(OutputFileName)
#---------------------------- END MESSAGE ---------------------------- 
def show_completion_popup(OutputFileName):
    def open_file():
        try:
            os.startfile(OutputFileName)  # Windows
        except AttributeError:
            subprocess.call(["open", OutputFileName])  # macOS fallback
        popup.destroy()

    def close_popup():
        popup.destroy()

    popup = tk.Tk()
    popup.title("Purrrrrfect!")
    popup.geometry("500x250")
    popup.configure(bg="#C2CAE8")

    label = tk.Label(
        popup,
        text=f"New Cat Report, who dis!\n\nSaved to:\n{OutputFileName}",
        bg="#C2CAE8",
        fg="#000000",
        font=("Arial", 11),
        justify="center",
        wraplength=380
    )
    label.pack(pady=10)

    button_frame = tk.Frame(popup, bg="#C2CAE8")

    print("Packing button frame…")
    button_frame.pack(pady=5)

    tk.Button(button_frame, text="View File", command=open_file, bg="#CBCEDF", font=("Arial", 10)).pack(side="left", padx=10)
    tk.Button(button_frame, text="OK", command=close_popup, bg="#CBCEDF", font=("Arial", 10)).pack(side="right", padx=10)

    popup.eval('tk::PlaceWindow . center')
    popup.update_idletasks()
    popup.mainloop()

OutputFileName = os.path.abspath(OutputFileName)
show_completion_popup(OutputFileName)



